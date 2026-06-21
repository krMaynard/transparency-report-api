"""Harvest harmonized DSA transparency reports (EU 2024/2835).

Downloads the 11-part CSV/XLSX reports published by platforms under the DSA
harmonized template (Commission Implementing Regulation (EU) 2024/2835, in
force from 1 July 2025), maps them to the research-api star schema, and
inserts them into demo.db alongside existing VLOP data.

Each downloaded report gets its own row in the `reports` dimension table;
fact rows reference it via `report_id`, enabling multi-period, multi-tier
queries across the same fact tables used for VLOP data.

Template → DB table mapping (Annex I parts):
  Part 1  → reports dimension row + services upsert
  Part 2  → categories upsert
  Part 3  → t3_member_state_orders
  Part 4  → t4_notices
  Part 5  → t5_own_initiative_illegal
  Part 6  → t6_own_initiative_tos
  Part 7  → t7_appeals_recidivism
  Part 8  → t8_automated_means
  Part 9  → t9_human_resources
  Part 10 → t10_amar
  Part 11 → t11_qualitative  (qualitative template)

Usage:
    # Harvest all services with a non-null report_url:
    python scripts/harvest-harmonized.py

    # Single service, dry-run (download + parse, no DB writes):
    python scripts/harvest-harmonized.py --service YouTube --dry-run

    # Use a pre-downloaded directory of CSV parts (skips HTTP fetch):
    python scripts/harvest-harmonized.py --service YouTube \\
        --source-dir /tmp/youtube-h2-2025/

    # Point at a non-default registry or DB:
    python scripts/harvest-harmonized.py \\
        --registry data/registry.json --db demo.db

Column-name notes
-----------------
The exact header names in the official Annex I CSV/XLSX depend on the
language version and any corrigendum.  The mapping dicts below (_T3_COLS
etc.) use the English names from the Commission's published XLSX template.
If a platform uses different names, pass --column-map /path/to/map.json
(a JSON object {"canonical_name": "platform_specific_name", ...}).
Run with --show-headers to print the actual CSV headers without inserting.
"""
import argparse
import csv
import json
import os
import sqlite3
import sys
import tempfile
import urllib.request
import zipfile
from pathlib import Path
from typing import Any

HERE = Path(__file__).parent.parent
_DEFAULT_REGISTRY = HERE / "data" / "registry.json"
_DEFAULT_DB = os.getenv("DB_PATH", str(HERE / "demo.db"))

# ---------------------------------------------------------------------------
# Official Annex I column name mapping (English, Commission XLSX template).
# Left side = our canonical name used in SQL; right side = Annex I header.
# Update the right-hand values if actual files use different casing/spacing.
# ---------------------------------------------------------------------------

# Part 1 — Report Identification
_P1_COLS: dict[str, str] = {
    "service_name":   "Service identifier",
    "period_start":   "Reporting period start date",
    "period_end":     "Reporting period end date",
    "publication_date": "Date of publication",
    "tier":           "Type of provider",       # maps to our tier vocab below
}

# Part 2 — Category Names
_P2_COLS: dict[str, str] = {
    "category_code":  "Category code",
    "category_label": "Category name",
}

# Part 3 — Member-State Orders (Art. 9 & 10)
_T3_COLS: dict[str, str] = {
    "category_code":          "Category code",
    "scope":                  "Type of order",        # "Removal" | "Information"
    "orders_to_act":          "Number of removal orders",
    "items":                  "Number of items subject to removal orders",
    "orders_to_provide_info": "Number of information orders",
}

# Part 4 — Notices (Art. 16)
_T4_COLS: dict[str, str] = {
    "category_code":    "Category code",
    "notices":          "Number of notices received",
    "tf_notices":       "Number of notices from trusted flaggers",
    "items":            "Number of items subject to notices",
    "tf_items":         "Number of items subject to trusted-flagger notices",
    "median_time":      "Median time to process notices (days)",
    "tf_median_time":   "Median time to process trusted-flagger notices (days)",
    "actions_law":      "Actions taken on legal basis",
    "tf_actions_law":   "Trusted-flagger actions on legal basis",
    "actions_tos":      "Actions taken on terms-and-conditions basis",
    "tf_actions_tos":   "Trusted-flagger actions on T&C basis",
}

# Part 5 — Own-Initiative on Illegal Content
_T5_COLS: dict[str, str] = {
    "category_code":               "Category code",
    "measures":                    "Total measures",
    "automated":                   "Measures applied using automated means",
    "vis_removal":                 "Visibility: removal",
    "vis_disable":                 "Visibility: disabling access",
    "vis_demoted":                 "Visibility: demotion",
    "vis_age_restricted":          "Visibility: age restriction",
    "vis_interaction_restricted":  "Visibility: interaction restriction",
    "vis_labelled":                "Visibility: labelling",
    "vis_other":                   "Visibility: other",
    "monetary_suspension":         "Monetary: suspension",
    "monetary_termination":        "Monetary: termination",
    "monetary_other":              "Monetary: other",
    "service_suspension":          "Service: suspension",
    "service_termination":         "Service: termination",
    "account_suspension":          "Account: suspension",
    "account_termination":         "Account: termination",
}

# Part 6 — Own-Initiative on T&C (same measures as Part 5, plus surface)
_T6_COLS: dict[str, str] = {
    **_T5_COLS,
    "surface": "Surface type",
}

# Part 7 — Appeals & Recidivism
_T7_COLS: dict[str, str] = {
    "section":    "Section",
    "indicator":  "Indicator",
    "scope":      "Scope",
    "surface":    "Surface",
    "value":      "Value",
}

# Part 8 — Automated Means (same shape as Part 7)
_T8_COLS: dict[str, str] = dict(_T7_COLS)

# Part 9 — Human Resources (same shape minus surface)
_T9_COLS: dict[str, str] = {
    "section":   "Section",
    "indicator": "Indicator",
    "scope":     "Scope",
    "value":     "Value",
}

# Part 10 — AMAR
_T10_COLS: dict[str, str] = {
    "scope": "Scope",
    "value": "Average monthly active recipients",
}

# Part 11 — Qualitative
_T11_COLS: dict[str, str] = {
    "indicator":  "Indicator",
    "value_text": "Description",
}

# Tier strings in the Annex I template → our vocabulary
_TIER_MAP: dict[str, str] = {
    "very large online platform":       "vlop",
    "very large online search engine":  "vlose",
    "vlop":                             "vlop",
    "vlose":                            "vlose",
    "online platform":                  "online-platform",
    "hosting service":                  "hosting",
    "intermediary service":             "intermediary",
}

# Expected CSV filenames inside a ZIP (may vary by platform/language).
# Keys are the part number (3-11); values are candidate filename prefixes.
_PART_FILENAMES: dict[int, list[str]] = {
    1:  ["part1", "report_identification", "identification"],
    2:  ["part2", "categories", "category_names"],
    3:  ["part3", "member_state_orders", "orders"],
    4:  ["part4", "notices"],
    5:  ["part5", "own_initiative_illegal", "illegal_content"],
    6:  ["part6", "own_initiative_tos", "tos_violations"],
    7:  ["part7", "appeals_recidivism", "appeals"],
    8:  ["part8", "automated_means", "automation"],
    9:  ["part9", "human_resources"],
    10: ["part10", "amar", "active_recipients"],
    11: ["part11", "qualitative"],
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _open_csv(path: Path, col_map: dict[str, str]) -> tuple[list[str], list[dict[str, str]]]:
    """Read a CSV file; return (headers, rows-as-dicts-keyed-by-canonical-name)."""
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return [], []
        raw_headers = list(reader.fieldnames)
        # Invert col_map: official_header → canonical_name
        inv = {v.lower(): k for k, v in col_map.items()}
        rows = []
        for raw_row in reader:
            row = {inv.get(h.lower().strip(), h): v for h, v in raw_row.items() if h is not None}
            rows.append(row)
    return raw_headers, rows


def _find_part(source_dir: Path, part: int) -> Path | None:
    """Locate the CSV file for a given part number in source_dir.

    If no CSV is found but an XLSX file is present, its sheets are extracted
    into the same directory (once) and the search is retried.
    """
    candidates = _PART_FILENAMES.get(part, [f"part{part}"])
    for name in source_dir.iterdir():
        stem = name.stem.lower().replace("-", "_").replace(" ", "_")
        if any(stem.startswith(c) or c in stem for c in candidates):
            if name.suffix.lower() in (".csv", ".txt"):
                return name
    # Fallback: extract any XLSX found in the directory
    for xlsx in source_dir.glob("*.xlsx"):
        converted = _xlsx_to_csvs(xlsx, source_dir)
        if converted:
            return _find_part(source_dir, part)  # retry after extraction
    return None


def _xlsx_to_csvs(xlsx_path: Path, dest: Path) -> list[Path]:
    """Split an XLSX file into per-part CSV files in dest (part{N}.csv).

    Returns paths to the written CSVs.  Skips gracefully if openpyxl is absent.
    """
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        print(f"  WARN openpyxl not installed — skipping {xlsx_path.name}. "
              "Run: pip install openpyxl")
        return []
    import re as _re
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    written: list[Path] = []
    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        # Infer part number from sheet name leading digit or keyword
        sn = sheet_name.lower().replace("-", "_").replace(" ", "_")
        m = _re.match(r"^(\d+)", sn.strip())
        part_num = int(m.group(1)) if m and 1 <= int(m.group(1)) <= 11 else i
        out_csv = dest / f"part{part_num}.csv"
        if not out_csv.exists():
            with out_csv.open("w", newline="", encoding="utf-8") as f:
                import csv as _csv
                w = _csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow(["" if v is None else v for v in row])
            print(f"    XLSX '{sheet_name}' → {out_csv.name}")
        written.append(out_csv)
    return written


def _download(url: str, dest: Path) -> None:
    req = urllib.request.Request(url, headers={"User-Agent": "transparency-report-harvester/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp, dest.open("wb") as f:
        while chunk := resp.read(65536):
            f.write(chunk)


def _extract_zip(zip_path: Path, dest: Path) -> None:
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(dest)


def _int_or_none(val: str | None) -> int | None:
    if val is None or val.strip() in ("", "N/A", "n/a", "-"):
        return None
    try:
        return int(val.replace(",", "").replace(" ", ""))
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Dimension upserts
# ---------------------------------------------------------------------------

def _upsert_service(conn: sqlite3.Connection, name: str, platform: str) -> int:
    row = conn.execute("SELECT id FROM services WHERE name = ?", (name,)).fetchone()
    if row:
        return row[0]
    cur = conn.execute("SELECT COALESCE(MAX(id)+1, 0) FROM services")
    new_id = cur.fetchone()[0]
    conn.execute("INSERT INTO services (id, name, platform) VALUES (?, ?, ?)",
                 (new_id, name, platform))
    return new_id


def _upsert_category(conn: sqlite3.Connection, code: str, label: str) -> int:
    row = conn.execute("SELECT id FROM categories WHERE code = ?", (code,)).fetchone()
    if row:
        return row[0]
    cur = conn.execute("SELECT COALESCE(MAX(id)+1, 0) FROM categories")
    new_id = cur.fetchone()[0]
    conn.execute("INSERT INTO categories (id, code, label) VALUES (?, ?, ?)",
                 (new_id, code, label))
    return new_id


def _upsert_dim(conn: sqlite3.Connection, table: str, name: str) -> int:
    row = conn.execute(f"SELECT id FROM {table} WHERE name = ?", (name,)).fetchone()
    if row:
        return row[0]
    cur = conn.execute(f"SELECT COALESCE(MAX(id)+1, 0) FROM {table}")
    new_id = cur.fetchone()[0]
    conn.execute(f"INSERT INTO {table} (id, name) VALUES (?, ?)", (new_id, name))
    return new_id


def _insert_report(conn: sqlite3.Connection, service_name: str, period_start: str,
                   period_end: str, tier: str, generated: str | None) -> int:
    """Insert a new reports row; return its id.

    If a row already exists for the same period_start/period_end/tier, it is
    reused.  service_name is accepted for forward-compatibility (a future schema
    may add a service_id FK to reports).
    """
    existing = conn.execute(
        "SELECT id FROM reports WHERE period_start = ? AND period_end = ? AND tier = ?",
        (period_start, period_end, tier),
    ).fetchone()
    if existing:
        # Reuse the existing report row for this period/tier (e.g. multiple services
        # from the same reporting cycle share one reports row in the current schema).
        return existing[0]
    cur = conn.execute("SELECT COALESCE(MAX(id)+1, 0) FROM reports")
    new_id = cur.fetchone()[0]
    period = f"{period_start}/{period_end}"
    conn.execute(
        "INSERT INTO reports (id, period, period_start, period_end, tier, generated) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (new_id, period, period_start, period_end, tier, generated),
    )
    return new_id


# ---------------------------------------------------------------------------
# Per-part ingestion
# ---------------------------------------------------------------------------

def _ingest_t3(conn: sqlite3.Connection, rows: list[dict], report_id: int,
               service_id: int, col_map: dict[str, str]) -> int:
    inserted = 0
    for row in rows:
        cat_id = _upsert_category(conn, row.get("category_code", ""), row.get("category_code", ""))
        scope_raw = row.get("scope", "Total number")
        scope_id = _upsert_dim(conn, "scopes", scope_raw)
        conn.execute(
            "INSERT INTO t3_member_state_orders VALUES (?,?,?,?,?,?,?)",
            (report_id, service_id, cat_id, scope_id,
             _int_or_none(row.get("orders_to_act")),
             _int_or_none(row.get("items")),
             _int_or_none(row.get("orders_to_provide_info"))),
        )
        inserted += 1
    return inserted


def _ingest_t4(conn: sqlite3.Connection, rows: list[dict], report_id: int,
               service_id: int, col_map: dict[str, str]) -> int:
    inserted = 0
    for row in rows:
        cat_id = _upsert_category(conn, row.get("category_code", ""), row.get("category_code", ""))
        conn.execute(
            "INSERT INTO t4_notices VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (report_id, service_id, cat_id,
             _int_or_none(row.get("notices")),
             _int_or_none(row.get("tf_notices")),
             _int_or_none(row.get("items")),
             _int_or_none(row.get("tf_items")),
             _int_or_none(row.get("median_time")),
             _int_or_none(row.get("tf_median_time")),
             _int_or_none(row.get("actions_law")),
             _int_or_none(row.get("tf_actions_law")),
             _int_or_none(row.get("actions_tos")),
             _int_or_none(row.get("tf_actions_tos"))),
        )
        inserted += 1
    return inserted


def _ingest_t5(conn: sqlite3.Connection, rows: list[dict], report_id: int,
               service_id: int, col_map: dict[str, str]) -> int:
    inserted = 0
    for row in rows:
        cat_id = _upsert_category(conn, row.get("category_code", ""), row.get("category_code", ""))
        conn.execute(
            "INSERT INTO t5_own_initiative_illegal VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (report_id, service_id, cat_id,
             _int_or_none(row.get("measures")),
             _int_or_none(row.get("automated")),
             _int_or_none(row.get("vis_removal")),
             _int_or_none(row.get("vis_disable")),
             _int_or_none(row.get("vis_demoted")),
             _int_or_none(row.get("vis_age_restricted")),
             _int_or_none(row.get("vis_interaction_restricted")),
             _int_or_none(row.get("vis_labelled")),
             _int_or_none(row.get("vis_other")),
             _int_or_none(row.get("monetary_suspension")),
             _int_or_none(row.get("monetary_termination")),
             _int_or_none(row.get("monetary_other")),
             _int_or_none(row.get("service_suspension")),
             _int_or_none(row.get("service_termination")),
             _int_or_none(row.get("account_suspension")),
             _int_or_none(row.get("account_termination"))),
        )
        inserted += 1
    return inserted


def _ingest_t6(conn: sqlite3.Connection, rows: list[dict], report_id: int,
               service_id: int, col_map: dict[str, str]) -> int:
    inserted = 0
    for row in rows:
        cat_id = _upsert_category(conn, row.get("category_code", ""), row.get("category_code", ""))
        surface_id = _upsert_dim(conn, "surfaces", row.get("surface", "All"))
        conn.execute(
            "INSERT INTO t6_own_initiative_tos VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (report_id, service_id, cat_id,
             _int_or_none(row.get("measures")),
             _int_or_none(row.get("automated")),
             _int_or_none(row.get("vis_removal")),
             _int_or_none(row.get("vis_disable")),
             _int_or_none(row.get("vis_demoted")),
             _int_or_none(row.get("vis_age_restricted")),
             _int_or_none(row.get("vis_interaction_restricted")),
             _int_or_none(row.get("vis_labelled")),
             _int_or_none(row.get("vis_other")),
             _int_or_none(row.get("monetary_suspension")),
             _int_or_none(row.get("monetary_termination")),
             _int_or_none(row.get("monetary_other")),
             _int_or_none(row.get("service_suspension")),
             _int_or_none(row.get("service_termination")),
             _int_or_none(row.get("account_suspension")),
             _int_or_none(row.get("account_termination")),
             surface_id),
        )
        inserted += 1
    return inserted


def _ingest_t7_t8(conn: sqlite3.Connection, table: str, rows: list[dict],
                  report_id: int, service_id: int) -> int:
    inserted = 0
    for row in rows:
        sec_id  = _upsert_dim(conn, "sections",   row.get("section", ""))
        ind_id  = _upsert_dim(conn, "indicators", row.get("indicator", ""))
        scp_id  = _upsert_dim(conn, "scopes",     row.get("scope", "Total number"))
        surf_id = _upsert_dim(conn, "surfaces",   row.get("surface", "All"))
        conn.execute(
            f"INSERT INTO {table} VALUES (?,?,?,?,?,?,?)",
            (report_id, service_id, sec_id, ind_id, scp_id,
             _int_or_none(row.get("value")), surf_id),
        )
        inserted += 1
    return inserted


def _ingest_t9(conn: sqlite3.Connection, rows: list[dict], report_id: int,
               service_id: int) -> int:
    inserted = 0
    for row in rows:
        sec_id = _upsert_dim(conn, "sections",   row.get("section", ""))
        ind_id = _upsert_dim(conn, "indicators", row.get("indicator", ""))
        scp_id = _upsert_dim(conn, "scopes",     row.get("scope", "Total number"))
        conn.execute(
            "INSERT INTO t9_human_resources VALUES (?,?,?,?,?,?)",
            (report_id, service_id, sec_id, ind_id, scp_id,
             _int_or_none(row.get("value"))),
        )
        inserted += 1
    return inserted


def _ingest_t10(conn: sqlite3.Connection, rows: list[dict], report_id: int,
                service_id: int) -> int:
    inserted = 0
    for row in rows:
        scp_id = _upsert_dim(conn, "scopes", row.get("scope", "Total number"))
        conn.execute(
            "INSERT INTO t10_amar VALUES (?,?,?,?)",
            (report_id, service_id, scp_id, _int_or_none(row.get("value"))),
        )
        inserted += 1
    return inserted


def _ingest_t11(conn: sqlite3.Connection, rows: list[dict], report_id: int,
                service_id: int) -> int:
    inserted = 0
    for row in rows:
        ind_id = _upsert_dim(conn, "indicators", row.get("indicator", ""))
        conn.execute(
            "INSERT INTO t11_qualitative VALUES (?,?,?,?)",
            (report_id, service_id, ind_id, row.get("value_text")),
        )
        inserted += 1
    return inserted


# ---------------------------------------------------------------------------
# Per-service harvest
# ---------------------------------------------------------------------------

def harvest_service(entry: dict[str, Any], db_path: str,
                    source_dir: Path | None, col_overrides: dict[str, str],
                    dry_run: bool, show_headers: bool) -> dict[str, int]:
    """Download + parse one service's report; return per-table insert counts."""
    name     = entry["service_name"]
    platform = entry["platform"]
    tier_raw = entry.get("tier", "vlop")
    period_start = entry["period_start"]
    period_end   = entry["period_end"]
    report_url   = entry.get("report_url")

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Harvesting {name} ({tier_raw}) "
          f"{period_start}–{period_end}")

    # Determine source directory
    work_dir: Path | None = source_dir
    _tmpdir = None
    if work_dir is None:
        if not report_url:
            print("  SKIP — no report_url in registry")
            return {}
        _tmpdir = tempfile.TemporaryDirectory(prefix=f"harvest_{name.replace(' ', '_')}_")
        work_dir = Path(_tmpdir.name)
        print(f"  Downloading {report_url}")
        dl_path = work_dir / "report.zip"
        _download(report_url, dl_path)
        if zipfile.is_zipfile(dl_path):
            _extract_zip(dl_path, work_dir)
        else:
            dl_path.rename(work_dir / "report.csv")

    if show_headers:
        print(f"  Files in {work_dir}:")
        for f in sorted(work_dir.iterdir()):
            print(f"    {f.name}")
            if f.suffix.lower() == ".csv":
                with f.open(encoding="utf-8-sig") as fh:
                    headers = next(csv.reader(fh), [])
                print(f"      columns: {headers}")
        return {}

    # Build merged column maps (base + overrides)
    col_map_by_part: dict[int, dict[str, str]] = {
        1: {**_P1_COLS, **col_overrides},
        2: {**_P2_COLS, **col_overrides},
        3: {**_T3_COLS, **col_overrides},
        4: {**_T4_COLS, **col_overrides},
        5: {**_T5_COLS, **col_overrides},
        6: {**_T6_COLS, **col_overrides},
        7: {**_T7_COLS, **col_overrides},
        8: {**_T8_COLS, **col_overrides},
        9: {**_T9_COLS, **col_overrides},
        10: {**_T10_COLS, **col_overrides},
        11: {**_T11_COLS, **col_overrides},
    }

    # Parse each part
    parsed: dict[int, list[dict]] = {}
    for part_num in range(1, 12):
        path = _find_part(work_dir, part_num)
        if path is None:
            print(f"  WARN  Part {part_num}: not found in {work_dir}")
            parsed[part_num] = []
            continue
        _, rows = _open_csv(path, col_map_by_part[part_num])
        parsed[part_num] = rows
        print(f"  Part {part_num}: {path.name} → {len(rows)} rows")

    if dry_run:
        return {f"part{p}": len(r) for p, r in parsed.items()}

    # Determine tier from Part 1 (override registry if present)
    p1 = parsed.get(1, [{}])
    p1_row = p1[0] if p1 else {}
    tier_from_file = _TIER_MAP.get(p1_row.get("tier", "").lower(), tier_raw)
    generated = p1_row.get("publication_date")

    # Upsert categories from Part 2
    conn = sqlite3.connect(db_path)
    try:
        with conn:
            service_id = _upsert_service(conn, name, platform)
            for row in parsed.get(2, []):
                _upsert_category(conn, row.get("category_code", ""), row.get("category_label", ""))

            report_id = _insert_report(conn, name, period_start, period_end,
                                       tier_from_file, generated)
            counts: dict[str, int] = {}
            counts["t3"] = _ingest_t3(conn, parsed.get(3, []), report_id, service_id, col_overrides)
            counts["t4"] = _ingest_t4(conn, parsed.get(4, []), report_id, service_id, col_overrides)
            counts["t5"] = _ingest_t5(conn, parsed.get(5, []), report_id, service_id, col_overrides)
            counts["t6"] = _ingest_t6(conn, parsed.get(6, []), report_id, service_id, col_overrides)
            counts["t7"] = _ingest_t7_t8(conn, "t7_appeals_recidivism",   parsed.get(7, []), report_id, service_id)
            counts["t8"] = _ingest_t7_t8(conn, "t8_automated_means",      parsed.get(8, []), report_id, service_id)
            counts["t9"] = _ingest_t9(conn, parsed.get(9, []),  report_id, service_id)
            counts["t10"] = _ingest_t10(conn, parsed.get(10, []), report_id, service_id)
            counts["t11"] = _ingest_t11(conn, parsed.get(11, []), report_id, service_id)

        total = sum(counts.values())
        print(f"  Inserted {total} rows (report_id={report_id}, service_id={service_id})")
        for t, n in counts.items():
            if n:
                print(f"    {t}: {n}")
        return counts
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Harvest harmonized DSA transparency reports (EU 2024/2835) into demo.db"
    )
    parser.add_argument("--registry", default=str(_DEFAULT_REGISTRY),
                        help="Path to registry.json (default: data/registry.json)")
    parser.add_argument("--db", default=_DEFAULT_DB,
                        help="SQLite database path (default: $DB_PATH or demo.db)")
    parser.add_argument("--service", metavar="NAME",
                        help="Harvest only this service_name (case-sensitive)")
    parser.add_argument("--source-dir", metavar="DIR",
                        help="Use pre-downloaded CSV directory instead of fetching report_url")
    parser.add_argument("--column-map", metavar="FILE",
                        help="JSON file mapping canonical column names to platform-specific ones")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse files without writing to the DB")
    parser.add_argument("--show-headers", action="store_true",
                        help="Print CSV headers found in source-dir and exit")
    args = parser.parse_args()

    col_overrides: dict[str, str] = {}
    if args.column_map:
        with open(args.column_map, encoding="utf-8") as f:
            col_overrides = json.load(f)

    with open(args.registry, encoding="utf-8") as f:
        registry: list[dict[str, Any]] = json.load(f)

    if args.service:
        registry = [e for e in registry if e["service_name"] == args.service]
        if not registry:
            print(f"Service '{args.service}' not found in registry.", file=sys.stderr)
            sys.exit(1)

    source_dir = Path(args.source_dir) if args.source_dir else None

    totals: dict[str, int] = {}
    skipped = 0
    for entry in registry:
        if not args.source_dir and not entry.get("report_url"):
            skipped += 1
            continue
        counts = harvest_service(
            entry, args.db, source_dir, col_overrides,
            dry_run=args.dry_run, show_headers=args.show_headers,
        )
        for t, n in counts.items():
            totals[t] = totals.get(t, 0) + n

    if skipped:
        print(f"\n{skipped} services skipped (no report_url) — add URLs to registry.json")
    if totals:
        print(f"\nTotal rows inserted: {sum(totals.values())}")
        for t, n in sorted(totals.items()):
            print(f"  {t}: {n}")


if __name__ == "__main__":
    main()
