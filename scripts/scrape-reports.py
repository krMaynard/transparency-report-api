"""Browser-automated downloader for harmonized DSA transparency reports.

Visits each platform's transparency_page_url from data/registry.json,
navigates to the download section by executing the entry's click_steps,
intercepts the file download, and saves it to data/reports/<slug>/.

ZIP archives are extracted in-place. XLSX files are split into per-part
CSV files (one sheet = one CSV, named part{N}.csv) so that
scripts/harvest-harmonized.py can ingest them with --source-dir.

Requirements (install separately from the main app):
    pip install playwright openpyxl
    playwright install chromium

Usage:
    # Download all services that have a transparency_page_url:
    python scripts/scrape-reports.py

    # Single service (case-insensitive substring match):
    python scripts/scrape-reports.py --service Snapchat

    # Show what would happen without opening a browser:
    python scripts/scrape-reports.py --dry-run

    # Watch the browser window:
    python scripts/scrape-reports.py --service TikTok --headful

    # Override registry or output directory:
    python scripts/scrape-reports.py --registry data/registry.json --out data/reports

Registry entry fields used here:
    transparency_page_url  The stable human-readable page for this platform's
                           DSA reports (must be set for scraping to work).
    click_steps            Optional list of navigation steps to reach the
                           download section before links are visible.
                           Each step is one of:
                             {"text": "Download data"}   click by visible text
                             {"selector": "#btn-dl"}     click by CSS selector
                             {"wait": ".dl-section"}     wait for selector
                             {"url": "https://..."}      navigate directly
    report_url             If set, skips browser automation and downloads
                           this URL directly with urllib (fallback / override).
"""
import argparse
import csv
import json
import re
import sys
import time
import urllib.request
import zipfile
from pathlib import Path
from typing import Any

HERE = Path(__file__).parent.parent
_DEFAULT_REGISTRY = HERE / "data" / "registry.json"
_DEFAULT_OUT = HERE / "data" / "reports"

# Annex I sheet/file → part number (mirrors _PART_FILENAMES in harvest-harmonized.py)
_PART_KEYWORDS: dict[int, list[str]] = {
    1:  ["part1", "report_identification", "identification"],
    2:  ["part2", "categories", "category_names", "category_name"],
    3:  ["part3", "member_state", "orders"],
    4:  ["part4", "notices"],
    5:  ["part5", "own_initiative_illegal", "illegal_content"],
    6:  ["part6", "own_initiative_tos", "tos", "terms"],
    7:  ["part7", "appeals", "recidivism"],
    8:  ["part8", "automated_means", "automation"],
    9:  ["part9", "human_resources"],
    10: ["part10", "amar", "active_recipients", "monthly_active"],
    11: ["part11", "qualitative"],
}

# Text / selector patterns that suggest a download trigger
_DL_TEXT_PATTERNS = [
    "download", "télécharger", "herunterladen", "scarica",
    "annex", "annexe", "anlage", "annesso",
    "csv", "xlsx", "zip", "data",
]

# Cookie-consent button selectors (tried in order, first match wins)
_COOKIE_SELECTORS = [
    "#onetrust-accept-btn-handler",
    "#accept-all-cookies",
    "[id*='accept'][id*='cookie']",
    "[class*='accept-all']",
    "button[aria-label*='Accept']",
    "button[aria-label*='accept']",
]
_COOKIE_TEXTS = [
    "Accept all", "Accept All", "Accept cookies", "Accept Cookies",
    "I agree", "Agree", "OK", "Akzeptieren", "Accepter tout",
    "Alles akzeptieren", "Tout accepter", "Accetta tutto",
]


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _sheet_to_part(sheet_name: str) -> int | None:
    """Infer Annex I part number from an XLSX sheet name."""
    sn = sheet_name.lower().replace("-", "_").replace(" ", "_").strip()
    # Leading digit: "3__member_state_orders" → 3
    m = re.match(r"^(\d+)", sn)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 11:
            return n
    # Keyword scan
    for part_num, keywords in _PART_KEYWORDS.items():
        if any(kw in sn for kw in keywords):
            return part_num
    return None


def _xlsx_to_csvs(xlsx_path: Path, dest: Path) -> list[Path]:
    """Split an XLSX file into per-part CSV files in dest.

    Returns paths to the written CSVs.
    """
    try:
        import openpyxl
    except ImportError:
        print("  ERROR openpyxl not installed — cannot convert XLSX. "
              "Run: pip install openpyxl", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    written: list[Path] = []
    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        part = _sheet_to_part(sheet_name) or i
        out_csv = dest / f"part{part}.csv"
        with out_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])
        print(f"    XLSX sheet '{sheet_name}' → {out_csv.name}")
        written.append(out_csv)
    return written


def _extract_and_convert(dl_path: Path, dest: Path) -> None:
    """Extract a ZIP (and convert any XLSX inside it) or convert a bare XLSX."""
    suffix = dl_path.suffix.lower()
    if suffix == ".zip" or zipfile.is_zipfile(dl_path):
        with zipfile.ZipFile(dl_path) as zf:
            zf.extractall(dest)
        print(f"  Extracted {dl_path.name} → {dest}")
        # Convert any XLSX found inside the ZIP
        for xlsx in dest.glob("*.xlsx"):
            print(f"  Converting {xlsx.name}")
            _xlsx_to_csvs(xlsx, dest)
    elif suffix in (".xlsx", ".xls"):
        print(f"  Converting {dl_path.name}")
        _xlsx_to_csvs(dl_path, dest)
    elif suffix in (".csv", ".txt"):
        target = dest / dl_path.name
        dl_path.rename(target)


def _dismiss_cookie_banner(page: Any) -> None:
    """Attempt to dismiss cookie consent banners (best-effort, never raises)."""
    try:
        # Selector-based
        for sel in _COOKIE_SELECTORS:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=1000):
                    btn.click(timeout=2000)
                    time.sleep(0.5)
                    return
            except Exception:
                continue
        # Text-based
        for text in _COOKIE_TEXTS:
            try:
                btn = page.get_by_role("button", name=re.compile(text, re.IGNORECASE)).first
                if btn.is_visible(timeout=800):
                    btn.click(timeout=2000)
                    time.sleep(0.5)
                    return
            except Exception:
                continue
    except Exception:
        pass


def _execute_step(page: Any, step: dict[str, str]) -> None:
    """Execute one click_step entry."""
    if "url" in step:
        page.goto(step["url"], wait_until="domcontentloaded")
        page.wait_for_load_state("networkidle", timeout=15000)
    elif "wait" in step:
        page.locator(step["wait"]).first.wait_for(state="visible", timeout=15000)
    elif "selector" in step:
        page.locator(step["selector"]).first.click(timeout=10000)
        page.wait_for_load_state("networkidle", timeout=10000)
    elif "text" in step:
        page.get_by_text(step["text"], exact=False).first.click(timeout=10000)
        page.wait_for_load_state("networkidle", timeout=10000)
    else:
        print(f"  WARN unknown step: {step}")


def _find_download_links(page: Any) -> list[tuple[str, str]]:
    """Scan the current page for likely DSA-report download links.

    Returns list of (link_text, href) for .zip/.xlsx/.csv hrefs or links
    whose visible text matches download-related keywords.
    """
    results: list[tuple[str, str]] = []
    links = page.locator("a[href]").all()
    for link in links:
        try:
            href: str = link.get_attribute("href") or ""
            text: str = (link.inner_text() or "").strip().lower()
        except Exception:
            continue
        href_low = href.lower()
        # Direct file extension
        if any(href_low.endswith(ext) for ext in (".zip", ".xlsx", ".csv")):
            results.append((text, href))
            continue
        # Text hint + file extension anywhere in URL
        if any(p in text for p in _DL_TEXT_PATTERNS):
            if any(ext in href_low for ext in (".zip", ".xlsx", ".csv")):
                results.append((text, href))
    return results


def _download_url(url: str, dest: Path) -> None:
    req = urllib.request.Request(
        url, headers={"User-Agent": "transparency-report-harvester/1.0"}
    )
    with urllib.request.urlopen(req, timeout=120) as resp, dest.open("wb") as f:
        f.write(resp.read())


def scrape_service(
    entry: dict[str, Any],
    out_base: Path,
    *,
    headful: bool = False,
    dry_run: bool = False,
) -> bool:
    """Download and extract one service's DSA report files.

    Returns True if files were written, False on skip/error.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("ERROR playwright not installed. Run: pip install playwright && "
              "playwright install chromium", file=sys.stderr)
        sys.exit(1)

    name = entry["service_name"]
    page_url: str | None = entry.get("transparency_page_url")
    report_url: str | None = entry.get("report_url")
    click_steps: list[dict] = entry.get("click_steps") or []

    if not page_url and not report_url:
        print(f"  SKIP {name} — no transparency_page_url or report_url")
        return False

    dest = out_base / _slug(name)

    if dry_run:
        print(f"  [DRY RUN] {name}")
        print(f"    page_url:    {page_url or '(none)'}")
        print(f"    report_url:  {report_url or '(none)'}")
        if click_steps:
            for s in click_steps:
                print(f"    step:        {s}")
        print(f"    output dir:  {dest}")
        return False

    dest.mkdir(parents=True, exist_ok=True)

    # Fast path: direct URL, no browser needed
    if report_url and not page_url:
        dl_path = dest / Path(report_url.split("?")[0]).name or "report.zip"
        print(f"  Downloading {report_url}")
        _download_url(report_url, dl_path)
        _extract_and_convert(dl_path, dest)
        return True

    # Browser path
    print(f"  Opening {page_url}")
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=not headful)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        try:
            page.goto(page_url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=15000)
            _dismiss_cookie_banner(page)

            for step in click_steps:
                print(f"    step: {step}")
                _execute_step(page, step)

            # Collect download links from the current page
            links = _find_download_links(page)
            if not links:
                print(f"  WARN no download links found on {page.url}")
                print("        Try adding click_steps to the registry entry, or "
                      "run with --headful to inspect the page manually.")
                return False

            print(f"  Found {len(links)} download link(s)")
            for text, href in links:
                # Resolve relative URLs
                if href.startswith("/"):
                    from urllib.parse import urlparse
                    parsed = urlparse(page.url)
                    href = f"{parsed.scheme}://{parsed.netloc}{href}"
                elif not href.startswith("http"):
                    continue

                fname = Path(href.split("?")[0]).name or "report.zip"
                dl_path = dest / fname
                print(f"  Downloading '{text or fname}' from {href}")
                try:
                    # Use page's network context (inherits cookies/session)
                    response = context.request.get(href, timeout=120000)
                    if response.status >= 400:
                        print(f"  WARN HTTP {response.status} for {href}")
                        continue
                    dl_path.write_bytes(response.body())
                    _extract_and_convert(dl_path, dest)
                except Exception as exc:
                    print(f"  WARN download failed: {exc}")
        finally:
            context.close()
            browser.close()

    return True


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Download harmonized DSA transparency reports via browser automation"
    )
    parser.add_argument("--registry", default=str(_DEFAULT_REGISTRY))
    parser.add_argument("--out", default=str(_DEFAULT_OUT),
                        help="Output base directory (default: data/reports/)")
    parser.add_argument("--service", metavar="NAME",
                        help="Only scrape services whose name contains NAME (case-insensitive)")
    parser.add_argument("--headful", action="store_true",
                        help="Show the browser window")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print plan without opening a browser")
    args = parser.parse_args()

    with open(args.registry, encoding="utf-8") as f:
        registry: list[dict[str, Any]] = json.load(f)

    if args.service:
        registry = [e for e in registry
                    if args.service.lower() in e["service_name"].lower()]
        if not registry:
            print(f"No services matching '{args.service}' in registry.",
                  file=sys.stderr)
            sys.exit(1)

    out_base = Path(args.out)
    out_base.mkdir(parents=True, exist_ok=True)

    ok = skipped = errors = 0
    for entry in registry:
        name = entry["service_name"]
        print(f"\n{name}")
        has_url = entry.get("transparency_page_url") or entry.get("report_url")
        if not has_url:
            print("  SKIP — add transparency_page_url to registry.json")
            skipped += 1
            continue
        try:
            result = scrape_service(entry, out_base,
                                    headful=args.headful,
                                    dry_run=args.dry_run)
            if result:
                ok += 1
            else:
                skipped += 1
        except Exception as exc:
            print(f"  ERROR {exc}")
            errors += 1

    print(f"\nDone — {ok} downloaded, {skipped} skipped, {errors} errors")
    if not args.dry_run and ok:
        print("\nRun harvest for each service:")
        print("  python scripts/harvest-harmonized.py "
              "--source-dir data/reports/<slug> --service <name>")


if __name__ == "__main__":
    main()
